import pathlib
import sys
import time
import logging
from typing import Optional
from pathlib import Path
from dataclasses import dataclass, field
from openpyxl import load_workbook, workbook, worksheet

# type hints
WorkBook = workbook.workbook.Workbook
WorkSheet = worksheet
PosixPath = pathlib.PosixPath


def check_path_is_file(path):
    if not Path(path).is_file():
        raise FileExistsError(f"{path = } not a file!!")


class MyLog:
    def __init__(self, folder_name: Optional[str] = 'run_logs',
                 file_level: Optional[str] = 'info'):
        """
        Initialize a new instance
        :param folder_name:Name of the folder where logs are stored
        """
        logger = logging.getLogger()
        logger.setLevel(logging.DEBUG)

        # log format config
        fmt = "%(asctime)s - %(filename)s[line:%(lineno)d] " \
              "- %(levelname)s: %(message)s"
        formatter = logging.Formatter(fmt)

        # log file config
        log_folder = Path('.').joinpath(folder_name)
        self.__creat_log_folder(log_folder)
        _log_time = time.strftime('%Y%m%d%H%M', time.localtime(time.time()))
        log_file_name = f"{_log_time}.log"
        log_file_path = str(log_folder.joinpath(log_file_name))

        file_handler = logging.FileHandler(log_file_path, mode='w')
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(formatter)
        stream_handler = logging.StreamHandler(sys.stdout)
        stream_handler.setLevel(logging.INFO)
        stream_handler.setFormatter(formatter)

        logger.addHandler(stream_handler)
        logger.addHandler(file_handler)
        self.logger = logger

    @staticmethod
    def __creat_log_folder(log_folder: Path):
        if not log_folder.is_dir():
            log_folder.mkdir()


@dataclass
class ExcelObj:
    """Excel Object """
    orig_file_path: str
    new_file_name: str = field(init=False)

    def __post_init__(self):
        check_path_is_file(self.orig_file_path)
        _parent_path = Path(self.orig_file_path).parent
        _new_file_name = self.orig_file_path.replace('.xls', '_new.xls')

        self.wb: Optional[WorkBook] = None  # Excel WorkBook
        self.orig_sheet_name_list: list = list()
        self.orig_first_sheet_name: Optional[str] = None
        self.orig_first_sheet: WorkSheet = None  # Excel First Work Sheet Object

        self.new_file_path = str(_parent_path.joinpath(_new_file_name))
        self.new_sheet_names: list = list()  # All new sheet name List
        self.new_sheets: list = list()  # All new sheet object List


class ExcelSplit:
    def __init__(self, orig_file_path: str, split_numb: int = 8000):
        """
        Initialize a new instance
        :param orig_file_path:  Path of the original Excel file
        :param split_numb: Every split_numb records are cut into a new sheet
        """
        self.logger = MyLog().logger
        self.excel = ExcelObj(orig_file_path=orig_file_path)

        _orig_file: Path = Path(orig_file_path)
        self.logger.info(f"Input Origin Excel File Path: {orig_file_path}, "
                         f"Absolute Path:{str(_orig_file.absolute())}")
        _file_name = self.excel.orig_file_path.replace('.xls', '_new.xls')
        self.new_file_path: PosixPath = _orig_file.parent.joinpath(_file_name)
        self.new_sheet_prefix: str = 'data_'
        self.split_num: int = split_numb

        self.__open_file()

    def __open_file(self):
        wb: WorkBook = load_workbook(filename=self.excel.orig_file_path)
        orig_sheet_names = wb.sheetnames
        orig_first_sheet_name = orig_sheet_names[0]
        self.logger.info(f"{orig_sheet_names = }, {orig_first_sheet_name = }")
        self.excel.wb = wb
        self.excel.orig_sheet_name_list = orig_sheet_names
        self.excel.orig_first_sheet_name = orig_first_sheet_name
        self.excel.orig_first_sheet = wb[orig_first_sheet_name]

        orig_max_row = self.excel.orig_first_sheet.max_row
        max_sheet_numb = int(orig_max_row / self.split_num) + 1
        self.logger.info(f"{self.excel.orig_first_sheet.dimensions = }")
        self.logger.info(f"{self.excel.orig_first_sheet.max_row = }")

        _new_sheet_list = [f"{self.new_sheet_prefix}{i + 1}" for i in
                           range(max_sheet_numb)]
        self.excel.new_sheet_names = _new_sheet_list
        self.logger.info(f"{self.excel.new_sheet_names = }")

    def creat_all_new_sheets(self):
        for i in self.excel.new_sheet_names:
            sheet = self.excel.wb.create_sheet(i)
            self.excel.new_sheets.append(sheet)

        self.logger.info(f"{self.excel.new_sheets = }")
        self.logger.info(f"all sheet names list: {self.excel.wb.sheetnames}")

    def write_all_new_sheet_record(self):
        for ws_name in self.excel.new_sheet_names:
            ws_index = int(ws_name.replace(self.new_sheet_prefix, ''))
            start_row_numb = self.split_num * (ws_index - 1) + 1

            _records = self.get_orig_sheet_mult_rows(
                start_row_numb=start_row_numb, count=self.split_num)

            if _records:
                self.__write_one_new_sheet(ws_name=ws_name, start_row_numb=1,
                                           src_data=_records)
            else:
                self.logger.warning(f"not get {_records = }")

    def get_orig_sheet_mult_rows(self, start_row_numb: Optional[int] = 1,
                                 count: Optional[int] = 8000) -> tuple:
        """
        Get multiple records from the original sheet
        :param start_row_numb: Start reading row number, Must be greater than 0
        :param count: Number of records to be read
        :return: Results of records read
        """
        ws: WorkSheet = self.excel.orig_first_sheet
        if (not isinstance(start_row_numb, int)) or start_row_numb <= 0:
            self.logger.error(f"{start_row_numb = } must be greater than 0 ")
            raise ValueError(f"{start_row_numb = } must be greater than 0 ")

        if start_row_numb > ws.max_row:
            self.logger.warning(f"{start_row_numb = } > {ws.max_row = }")
            return tuple()

        if (not isinstance(count, int)) or count <= 0:
            self.logger.error(f"{count = } must be greater than 0 ")
            raise ValueError(f"{count = } must be greater than 0 ")

        max_row = start_row_numb + count - 1
        if max_row > ws.max_row:
            max_row = ws.max_row

        return ws[start_row_numb: max_row]

    def __write_one_new_sheet(self, ws_name, start_row_numb, src_data) -> None:
        """
        Write the records read from the original table in a new sheet
        :param ws_name: The name of the sheet to be written
        :param start_row_numb: Start write row number, Must be greater than 0
        :param src_data: Original record
        """
        self.logger.info(f"start: {len(src_data) = }")
        w_ws = self.excel.wb[ws_name]
        for index, origin_data in enumerate(src_data):
            w_row = index + start_row_numb
            r_one_cell = origin_data[0]
            r_one_value = r_one_cell.value
            self.logger.debug(f'read from {r_one_cell = }, {r_one_value = },'
                              f'write to {w_ws = }, {w_row = }')
            w_ws.cell(row=w_row, column=r_one_cell.column, value=r_one_value)

    def save_to_disk(self, new_file_path: Optional[str] = None):
        _path = Path(new_file_path) if new_file_path else self.new_file_path
        self.logger.info(f"Start save to File Path: {_path}, "
                         f"Absolute Path:{str(_path.absolute())}")
        self.excel.wb.save(_path)


if __name__ == '__main__':
    file_path = f'file_001.xlsx'
    a = ExcelSplit(file_path)
    a.creat_all_new_sheets()
    a.write_all_new_sheet_record()
    a.save_to_disk()
